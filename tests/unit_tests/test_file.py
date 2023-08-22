from pathlib import Path

from openpyxl import load_workbook

from app.core import db_flush
from app.tasks import task, fill_repos, is_modified, read_file, init_repos

from tests import conftest as c
from tests.fixtures import data as d
from tests.utils import compare_lists

FAKE_FILE_PATH = Path('tests/fixtures/Menu.xlsx')


def write_file(fname: str, edit: bool = False) -> None:
    wb = load_workbook(filename=fname)
    ws = wb['Лист1']
    if edit:
        ws['B1'] = 'Menu'
    else:
        ws['B1'] = 'Меню'
    wb.save(filename=fname)


def test_menu_file_exists() -> None:
    assert c.FILE_PATH.exists(), f'No such file: {c.FILE_PATH}'


def test_read_file() -> None:
    menus, _, _ = read_file(FAKE_FILE_PATH)
    assert menus == d.EXPECTED_MENU_FILE_CONTENT


def test_is_modified() -> None:
    assert not is_modified(FAKE_FILE_PATH)
    write_file(FAKE_FILE_PATH)
    assert is_modified(FAKE_FILE_PATH)


@c.pytest_mark_anyio
async def test_db_flush(dish: c.Response,
                        get_menu_repo: c.MenuRepository,
                        get_submenu_repo: c.SubmenuRepository,
                        get_dish_repo: c.DishRepository) -> None:
    assert await get_menu_repo.get_all() is not None
    assert await get_submenu_repo.get_all() is not None
    assert await get_dish_repo.get_all() is not None
    await db_flush(c.engine)
    assert await get_menu_repo.get_all() is None
    assert await get_submenu_repo.get_all() is None
    assert await get_dish_repo.get_all() is None


@c.pytest_mark_anyio
async def test_fill_repos(get_menu_service: c.MenuService,
                          get_submenu_service: c.SubmenuService,
                          get_dish_service: c.DishService) -> None:
    assert await get_menu_service.db.get_all() is None
    assert await get_menu_service.redis.get_all() is None
    menus, _, _ = read_file(FAKE_FILE_PATH)
    await fill_repos(menus, get_menu_service, get_submenu_service, get_dish_service)

    menus_db = await get_menu_service.db.get_all()
    assert menus_db is not None
    assert len(menus_db) == 2
    compare_lists(menus_db, await get_menu_service.redis.get_all())


    submenus_db = await get_submenu_service.db.get_all()
    assert submenus_db is not None
    assert len(submenus_db) == 4
    compare_lists(submenus_db, await get_submenu_service.redis.get_all())

    dishes_db = await get_dish_service.db.get_all()
    assert dishes_db is not None
    assert len(dishes_db) == 12
    compare_lists(dishes_db, await get_dish_service.redis.get_all())


@c.pytest_mark_anyio
async def test_init_repos(dish: c.Response,
                          get_test_session: c.AsyncSession,
                          get_test_redis: c.FakeRedis) -> None:
    await init_repos(get_test_session, FAKE_FILE_PATH, c.engine, get_test_redis)
